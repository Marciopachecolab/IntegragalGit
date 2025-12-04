"""
plate_viewer_demo.py
--------------------

Módulo independente de TESTE para visualização de placa 8x12
tanto na interface quanto em um arquivo Excel gerado.

- Gera uma placa de exemplo com amostras, CN, CP, CT alvo, CT RP
  e resultado categórico (Detectado, Nao Detectado, Inconclusivo, Invalido).
- Mostra a placa em uma janela (customtkinter se disponível; senão tkinter).
- Exporta a placa para um arquivo Excel 'placa_teste.xlsx' no diretório atual.

Uso:
    python plate_viewer_demo.py
"""

from __future__ import annotations

import random
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Tentativa de import do customtkinter; fallback para tkinter puro
# ---------------------------------------------------------------------------
try:
    import customtkinter as ctk

    USE_CTK = True
except ImportError:
    import tkinter as tk

    USE_CTK = False

# ---------------------------------------------------------------------------
# Excel (openpyxl)
# ---------------------------------------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Dados da placa
# ---------------------------------------------------------------------------

ROWS = "ABCDEFGH"
COLS = list(range(1, 13))


@dataclass
class WellResult:
    well: str  # ex: "A1"
    sample_code: str  # código da amostra (ou CN/CP)
    ct_target: Optional[float]  # CT do alvo principal
    ct_rp: Optional[float]  # CT do RP
    result: str  # "Detectado", "Nao Detectado", "Inconclusivo", "Invalido"
    is_control: bool = False  # True para CN/CP
    control_type: Optional[str] = None  # "CN", "CP", etc.


def _well_name(row: int, col: int) -> str:
    """Converte índices 0-based para nome de poço (A1..H12)."""
    return f"{ROWS[row]}{col+1}"


def generate_dummy_plate() -> List[WellResult]:
    """
    Gera uma placa de teste com dados fictícios, mas verossímeis.

    - Posição A1..H12 preenchida.
    - Alguns poços marcados como CN/CP.
    - CTs sorteados em faixas plausíveis.
    """
    resultados: List[WellResult] = []

    # Define posições de CN/CP (exemplo: H11 = CN, H12 = CP)
    cn_well = "H11"
    cp_well = "H12"

    for r in range(8):
        for c in range(12):
            well = _well_name(r, c)

            # Controles CN / CP
            if well == cn_well:
                resultados.append(
                    WellResult(
                        well=well,
                        sample_code="CN",
                        ct_target=_rnd_or_none((35.0, 40.0)),  # idealmente não detecta
                        ct_rp=_rnd_or_none((20.0, 30.0)),
                        result="Nao Detectado",
                        is_control=True,
                        control_type="CN",
                    )
                )
                continue

            if well == cp_well:
                resultados.append(
                    WellResult(
                        well=well,
                        sample_code="CP",
                        ct_target=_rnd_or_none((20.0, 30.0)),
                        ct_rp=_rnd_or_none((20.0, 30.0)),
                        result="Detectado",
                        is_control=True,
                        control_type="CP",
                    )
                )
                continue

            # Amostras comuns
            sample_code = f"42121{r}{c:02d}"

            # Sorteia um tipo de resultado
            tipo = random.choices(
                population=["Detectado", "Nao Detectado", "Inconclusivo", "Invalido"],
                weights=[0.5, 0.3, 0.1, 0.1],
                k=1,
            )[0]

            if tipo == "Detectado":
                ct_t = _rnd_or_none((15.0, 35.0))
                ct_r = _rnd_or_none((20.0, 30.0))
            elif tipo == "Nao Detectado":
                ct_t = None
                ct_r = _rnd_or_none((20.0, 30.0))
            elif tipo == "Inconclusivo":
                ct_t = _rnd_or_none((35.0, 40.0))
                ct_r = _rnd_or_none((20.0, 30.0))
            else:  # Invalido
                ct_t = None
                ct_r = None

            resultados.append(
                WellResult(
                    well=well,
                    sample_code=sample_code,
                    ct_target=ct_t,
                    ct_rp=ct_r,
                    result=tipo,
                    is_control=False,
                    control_type=None,
                )
            )

    return resultados


def _rnd_or_none(interval: Tuple[float, float]) -> Optional[float]:
    """Sorteia um float em interval ou None (com pequena chance)."""
    if random.random() < 0.05:
        return None
    return round(random.uniform(*interval), 1)


def results_to_matrix(results: List[WellResult]) -> Dict[Tuple[int, int], WellResult]:
    """
    Indexa os resultados por posição (linha, coluna) 0-based.

    :return: dicionário {(row, col): WellResult}
    """
    idx: Dict[str, WellResult] = {r.well.upper(): r for r in results}
    matrix: Dict[Tuple[int, int], WellResult] = {}
    for r in range(8):
        for c in range(12):
            well = _well_name(r, c).upper()
            matrix[(r, c)] = idx.get(
                well,
                WellResult(
                    well=well,
                    sample_code="",
                    ct_target=None,
                    ct_rp=None,
                    result="",
                    is_control=False,
                    control_type=None,
                ),
            )
    return matrix


# ---------------------------------------------------------------------------
# Interface gráfica
# ---------------------------------------------------------------------------


def show_plate_gui(results: List[WellResult]) -> None:
    """
    Abre uma janela exibindo a placa 8x12.

    Cada poço mostra:
    - Código da amostra (ou CN/CP)
    - CT alvo / CT RP
    - Resultado com cor de fundo:
        Detectado   -> preto
        Nao Detectado -> cinza claro
        Inconclusivo -> laranja
        Invalido    -> vermelho
    """
    matrix = results_to_matrix(results)

    if USE_CTK:
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        root = ctk.CTk()
        root.title("Visualização de Placa - Demo")
        # janela ~40% mais larga
        root.geometry("1120x600")
    else:
        root = tk.Tk()
        root.title("Visualização de Placa - Demo")
        # janela ~40% mais larga
        root.geometry("1120x600")

    # Cabeçalhos de coluna
    for c in range(12):
        label = _make_label(root, text=str(c + 1), bold=True)
        label.grid(row=0, column=c + 1, padx=2, pady=2, sticky="nsew")

    # Cabeçalhos de linha + células
    for r in range(8):
        # Cabeçalho da linha (A..H)
        row_label = _make_label(root, text=ROWS[r], bold=True)
        row_label.grid(row=r + 1, column=0, padx=2, pady=2, sticky="nsew")

        for c in range(12):
            well_res = matrix[(r, c)]
            frame = _make_well_frame(root, well_res)
            frame.grid(row=r + 1, column=c + 1, padx=2, pady=2, sticky="nsew")

    # Ajustar pesos das colunas/linhas para expandir
    for c in range(13):
        root.grid_columnconfigure(c, weight=1)
    for r in range(9):
        root.grid_rowconfigure(r, weight=1)

    root.mainloop()


def _make_label(parent: Any, text: str, bold: bool = False):
    """Cria label simples para cabeçalhos, compatível com CTk ou Tk."""
    if USE_CTK:
        font = ("Arial", 11, "bold" if bold else "normal")
        return ctk.CTkLabel(parent, text=text, font=font)
    else:
        fnt = ("Arial", 11, "bold" if bold else "normal")
        return tk.Label(parent, text=text, font=fnt)


def _make_well_frame(parent: Any, well: WellResult):
    """Cria um 'quadradinho' de poço com informações de resultado."""
    # Cores básicas por resultado
    bg_main, fg_main = "white", "black"
    side_bg, side_fg = "#DDDDDD", "black"

    if well.result == "Detectado":
        side_bg, side_fg = "black", "white"
    elif well.result == "Nao Detectado":
        side_bg, side_fg = "#DDDDDD", "black"
    elif well.result == "Inconclusivo":
        side_bg, side_fg = "orange", "black"
    elif well.result == "Invalido":
        side_bg, side_fg = "red", "white"

    # Destacar CN/CP
    if well.is_control:
        bg_main = "#E0F2FF"

    if USE_CTK:
        frame = ctk.CTkFrame(parent, fg_color=bg_main, corner_radius=4)
        inner_font = ("Arial", 9)
        small_font = ("Arial", 8)
        side = ctk.CTkFrame(frame, width=10, fg_color=side_bg, corner_radius=2)
        side.pack(side="left", fill="y")

        side_label = ctk.CTkLabel(
            side,
            text=_short_result(well.result),
            font=small_font,
            text_color=side_fg,
        )
        side_label.pack(expand=True)

        code = ctk.CTkLabel(frame, text=well.sample_code, font=inner_font)
        code.pack(side="top", fill="x", padx=2, pady=(2, 1))

        ct_text = f"CT: {well.ct_target if well.ct_target is not None else '--'}"
        rp_text = f"RP: {well.ct_rp if well.ct_rp is not None else '--'}"

        ct_label = ctk.CTkLabel(frame, text=ct_text, font=small_font)
        ct_label.pack(side="top", fill="x")

        rp_label = ctk.CTkLabel(frame, text=rp_text, font=small_font)
        rp_label.pack(side="top", fill="x", pady=(0, 2))

    else:
        frame = tk.Frame(parent, bg=bg_main, bd=1, relief="solid")
        inner_font = ("Arial", 8)
        small_font = ("Arial", 7)

        side = tk.Frame(frame, width=10, bg=side_bg)
        side.pack(side="left", fill="y")

        side_label = tk.Label(
            side,
            text=_short_result(well.result),
            font=small_font,
            fg=side_fg,
            bg=side_bg,
        )
        side_label.pack(expand=True)

        code = tk.Label(frame, text=well.sample_code, font=inner_font, bg=bg_main)
        code.pack(side="top", fill="x", padx=2, pady=(2, 1))

        ct_text = f"CT: {well.ct_target if well.ct_target is not None else '--'}"
        rp_text = f"RP: {well.ct_rp if well.ct_rp is not None else '--'}"

        ct_label = tk.Label(frame, text=ct_text, font=small_font, bg=bg_main)
        ct_label.pack(side="top", fill="x")

        rp_label = tk.Label(frame, text=rp_text, font=small_font, bg=bg_main)
        rp_label.pack(side="top", fill="x", pady=(0, 2))

    return frame


def _short_result(result: str) -> str:
    """Converte resultado em abreviação para a barrinha lateral."""
    mapping = {
        "Detectado": "DET",
        "Nao Detectado": "ND",
        "Inconclusivo": "INC",
        "Invalido": "INV",
        "": "",
    }
    return mapping.get(result, result[:3].upper())


# ---------------------------------------------------------------------------
# Exportação para Excel
# ---------------------------------------------------------------------------


def export_plate_to_excel(results: List[WellResult], path: Path) -> None:
    """
    Gera um arquivo Excel com a placa 8x12.

    - Cabeçalhos de coluna (1..12) e linha (A..H).
    - Cada poço com:
        código da amostra,
        CT alvo,
        CT RP,
        resultado com cor de fundo semelhante ao modelo.
    """
    if not HAS_OPENPYXL:
        print("openpyxl não está instalado; pulei a geração do Excel.")
        return

    matrix = results_to_matrix(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "Placa"

    # Estilos básicos
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Cores de resultado (ajustadas)
    fill_det = PatternFill("solid", fgColor="FF9999")  # vermelho claro (Detectado)
    fill_nd = PatternFill("solid", fgColor="DDDDDD")  # cinza claro (Nao Detectado)
    fill_inc = PatternFill("solid", fgColor="ADD8E6")  # azul claro (Inconclusivo)
    fill_inv = PatternFill("solid", fgColor="C6EFCE")  # verde claro (Invalido)
    fill_ctrl = PatternFill("solid", fgColor="E0F2FF")  # azul claro para controles

    # Cabeçalhos
    ws["A1"] = ""
    for c in range(12):
        cell = ws.cell(row=1, column=c + 2)
        cell.value = c + 1
        cell.font = header_font
        cell.alignment = center

    for r in range(8):
        cell = ws.cell(row=r + 2, column=1)
        cell.value = ROWS[r]
        cell.font = header_font
        cell.alignment = center

    # Preenchimento dos poços
    for r in range(8):
        for c in range(12):
            well = matrix[(r, c)]
            row_excel = r + 2
            col_excel = c + 2

            cell = ws.cell(row=row_excel, column=col_excel)

            # Texto principal: código + CTs em múltiplas linhas
            lines = [
                well.sample_code,
                f"CT: {well.ct_target if well.ct_target is not None else '--'}",
                f"RP: {well.ct_rp if well.ct_rp is not None else '--'}",
                _short_result(well.result),
            ]
            cell.value = "\n".join(lines)
            cell.alignment = center
            cell.border = border_thin
            # fonte maior e em negrito (inclui o código da amostra)
            cell.font = Font(size=9, bold=True)

            # Cor de fundo
            fill = None
            if well.result == "Detectado":
                fill = fill_det
            elif well.result == "Nao Detectado":
                fill = fill_nd
            elif well.result == "Inconclusivo":
                fill = fill_inc
            elif well.result == "Invalido":
                fill = fill_inv

            if fill:
                cell.fill = fill

            if well.is_control:
                # destaca controles com azul claro
                cell.fill = fill_ctrl

    # Ajustar largura de colunas e altura de linhas
    ws.column_dimensions["A"].width = 4
    for col_letter in "BCDEFGHIJKLM":
        ws.column_dimensions[col_letter].width = 14

    for row_idx in range(2, 10):
        ws.row_dimensions[row_idx].height = 45

    # Salvar
    path = Path(path)
    wb.save(path)
    print(f"Arquivo Excel gerado em: {path.resolve()}")


# ---------------------------------------------------------------------------
# Execução direta
# ---------------------------------------------------------------------------


def main():
    random.seed(42)  # determinístico para teste
    results = generate_dummy_plate()

    # 1) Exporta Excel
    export_plate_to_excel(results, Path("placa_teste.xlsx"))

    # 2) Mostra GUI
    show_plate_gui(results)


if __name__ == "__main__":
    main()
