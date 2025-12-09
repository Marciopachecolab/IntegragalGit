"""
Análise detalhada da planilha Biomanguinhos VR1-VR2 PLACA 5
Para criar padrão customizado no detector
"""

import sys
from pathlib import Path
sys.path.insert(0, '.')

from services.equipment_detector import analisar_estrutura_xlsx
from openpyxl import load_workbook

planilha = "C:/Users/marci/Downloads/18 JULHO 2025/20250718 VR1-VR2 BIOM PLACA 5.xlsx"

print("="*80)
print("ANÁLISE DETALHADA - Planilha Biomanguinhos")
print("="*80)
print(f"\nArquivo: {Path(planilha).name}\n")

# Análise estrutural
print("📊 ESTRUTURA DETECTADA:")
print("-"*80)
estrutura = analisar_estrutura_xlsx(planilha)

print(f"\n1. Informações Gerais:")
print(f"   Total de colunas: {estrutura['total_colunas']}")
print(f"   Total de linhas com dados: {estrutura['total_linhas_dados']}")
print(f"   Linha início dos dados: {estrutura['linha_inicio_dados']}")
print(f"   Colunas não vazias: {len(estrutura['colunas_nao_vazias'])} colunas")

print(f"\n2. Headers Completos (primeiros 15):")
for i, header in enumerate(estrutura['headers'][:15]):
    letra_col = chr(65 + i) if i < 26 else f"A{chr(65 + i - 26)}"
    print(f"   [{letra_col}] Col {i:2d}: {header[:50]}")

print(f"\n3. Colunas Detectadas Automaticamente:")
print(f"   Well:   Coluna {estrutura['coluna_well']} ({chr(65 + estrutura['coluna_well']) if estrutura['coluna_well'] is not None else 'N/A'})")
print(f"   Sample: Coluna {estrutura['coluna_sample']} ({chr(65 + estrutura['coluna_sample']) if estrutura['coluna_sample'] is not None else 'N/A'})")
print(f"   Target: Coluna {estrutura['coluna_target']} ({chr(65 + estrutura['coluna_target']) if estrutura['coluna_target'] is not None else 'N/A'})")
print(f"   CT:     Coluna {estrutura['coluna_ct']} ({chr(65 + estrutura['coluna_ct']) if estrutura['coluna_ct'] is not None else 'N/A'})")

if estrutura['amostras_wells']:
    print(f"\n4. Amostras de valores na coluna Well:")
    for i, well in enumerate(estrutura['amostras_wells'][:10], 1):
        print(f"   {i}. {well}")

# Análise mais profunda com openpyxl
print("\n" + "="*80)
print("📋 ANÁLISE PROFUNDA (primeiras 20 linhas)")
print("="*80)

wb = load_workbook(planilha, read_only=True, data_only=True)
ws = wb.active

print(f"\nSheet ativa: {ws.title}")
print(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")

print("\n5. Primeiras 20 linhas (com valores):")
for row_idx in range(1, min(21, ws.max_row + 1)):
    row_values = []
    for col_idx in range(1, min(16, ws.max_column + 1)):  # Primeiras 15 colunas
        cell = ws.cell(row_idx, col_idx)
        val = str(cell.value) if cell.value is not None else ""
        if len(val) > 20:
            val = val[:17] + "..."
        row_values.append(val)
    
    # Mostrar apenas se tiver conteúdo
    if any(v.strip() for v in row_values):
        valores_str = " | ".join(f"{v:20s}" for v in row_values)
        print(f"   Linha {row_idx:2d}: {valores_str}")

wb.close()

# Propor padrão customizado
print("\n" + "="*80)
print("💡 PADRÃO CUSTOMIZADO SUGERIDO")
print("="*80)

padrao_sugerido = f"""
Equipamento: Biomanguinhos_VR
Headers esperados: {', '.join(estrutura['headers'][:5])}
Coluna Well: {estrutura['coluna_well']} (col {chr(65 + estrutura['coluna_well']) if estrutura['coluna_well'] is not None else 'N/A'})
Coluna Target: {estrutura['coluna_target']} (col {chr(65 + estrutura['coluna_target']) if estrutura['coluna_target'] is not None else 'N/A'})
Coluna CT: {estrutura['coluna_ct']} (col {chr(65 + estrutura['coluna_ct']) if estrutura['coluna_ct'] is not None else 'N/A'})
Linha início: {estrutura['linha_inicio_dados']}

Características:
- Total de linhas: ~{estrutura['total_linhas_dados']}
- Formato de well: {estrutura['amostras_wells'][0] if estrutura['amostras_wells'] else 'desconhecido'}
- Tipo de placa: 96 wells
"""

print(padrao_sugerido)

print("\n" + "="*80)
print("✅ Análise concluída!")
print("="*80)
