"""
Exportação de Relatórios - IntegaGal
Fase 3.4 - Interface Gráfica

Módulo para exportação de relatórios em múltiplos formatos (PDF, Excel, CSV)
"""

import os
from datetime import datetime
from typing import Dict, Any, Optional, List
import pandas as pd

# PDF
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


class ExportadorRelatorios:
    """
    Classe principal para exportação de relatórios
    Suporta PDF, Excel e CSV
    """
    
    def __init__(self, pasta_saida: str = "reports"):
        """
        Inicializa exportador
        
        Args:
            pasta_saida: Pasta onde os relatórios serão salvos
        """
        self.pasta_saida = pasta_saida
        
        # Criar pasta se não existir
        if not os.path.exists(pasta_saida):
            os.makedirs(pasta_saida)
    
    def exportar_exame_pdf(self, dados_exame: Dict[str, Any], nome_arquivo: Optional[str] = None) -> str:
        """
        Exporta detalhes de um exame para PDF
        
        Args:
            dados_exame: Dicionário com dados do exame
            nome_arquivo: Nome do arquivo (opcional, gera automaticamente)
        
        Returns:
            Caminho completo do arquivo gerado
        """
        # Nome do arquivo
        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            exame_nome = dados_exame.get('exame', 'exame').replace(' ', '_')
            nome_arquivo = f"relatorio_exame_{exame_nome}_{timestamp}.pdf"
        
        caminho_arquivo = os.path.join(self.pasta_saida, nome_arquivo)
        
        # Criar documento
        doc = SimpleDocTemplate(
            caminho_arquivo,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # Container para elementos
        elementos = []
        
        # Estilos
        estilos = getSampleStyleSheet()
        estilo_titulo = ParagraphStyle(
            'CustomTitle',
            parent=estilos['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1E88E5'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        estilo_subtitulo = ParagraphStyle(
            'CustomSubtitle',
            parent=estilos['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1E88E5'),
            spaceAfter=12,
            spaceBefore=20
        )
        estilo_normal = estilos['Normal']
        
        # Título
        titulo = Paragraph(f"<b>Relatório de Análise - {dados_exame.get('exame', 'N/A')}</b>", estilo_titulo)
        elementos.append(titulo)
        elementos.append(Spacer(1, 0.5*cm))
        
        # Informações gerais
        info_data = [
            ['Data/Hora:', dados_exame.get('data_hora', 'N/A')],
            ['Equipamento:', dados_exame.get('equipamento', 'N/A')],
            ['Status:', dados_exame.get('status', 'N/A').upper()],
        ]
        
        if 'analista' in dados_exame:
            info_data.append(['Analista:', dados_exame['analista']])
        
        tabela_info = Table(info_data, colWidths=[4*cm, 12*cm])
        tabela_info.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1E88E5')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elementos.append(tabela_info)
        elementos.append(Spacer(1, 1*cm))
        
        # Seção de Alvos
        alvos = dados_exame.get('alvos', {})
        if alvos:
            elementos.append(Paragraph("<b>Alvos Detectados</b>", estilo_subtitulo))
            
            alvos_data = [['Alvo', 'CT', 'Resultado', 'Status']]
            for nome_alvo, dados in alvos.items():
                ct = dados.get('ct', 'N/D')
                if isinstance(ct, (int, float)):
                    ct = f"{ct:.2f}"
                
                resultado = dados.get('resultado', 'N/D')
                status = '✓' if resultado in ('Detectado', 'Positivo') else '✗'
                
                alvos_data.append([nome_alvo, str(ct), resultado, status])
            
            tabela_alvos = Table(alvos_data, colWidths=[4*cm, 3*cm, 5*cm, 3*cm])
            tabela_alvos.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elementos.append(tabela_alvos)
            elementos.append(Spacer(1, 1*cm))
        
        # Seção de Controles
        controles = dados_exame.get('controles', {})
        if controles:
            elementos.append(Paragraph("<b>Controles de Qualidade</b>", estilo_subtitulo))
            
            controles_data = [['Controle', 'Tipo', 'CT', 'Status']]
            for nome_controle, dados in controles.items():
                tipo = dados.get('tipo', 'N/A')
                ct = dados.get('ct', 'N/D')
                if isinstance(ct, (int, float)):
                    ct = f"{ct:.2f}"
                
                status = dados.get('status', 'N/A')
                
                controles_data.append([nome_controle, tipo, str(ct), status])
            
            tabela_controles = Table(controles_data, colWidths=[5*cm, 3*cm, 3*cm, 4*cm])
            tabela_controles.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elementos.append(tabela_controles)
            elementos.append(Spacer(1, 1*cm))
        
        # Seção de Regras
        regras_resultado = dados_exame.get('regras_resultado')
        if regras_resultado:
            elementos.append(Paragraph("<b>Regras Aplicadas</b>", estilo_subtitulo))
            
            # Resumo
            detalhes = regras_resultado.get('detalhes', 'N/A')
            elementos.append(Paragraph(f"Resumo: {detalhes}", estilo_normal))
            elementos.append(Spacer(1, 0.5*cm))
            
            # Validações
            validacoes = regras_resultado.get('validacoes', [])
            if validacoes:
                regras_data = [['Regra', 'Resultado', 'Impacto', 'Detalhes']]
                for validacao in validacoes:
                    nome = validacao.get('regra_nome', 'N/A')
                    resultado = validacao.get('resultado', 'N/A')
                    impacto = validacao.get('impacto', 'N/A').upper()
                    detalhes = validacao.get('detalhes', '')
                    
                    # Truncar detalhes se muito longo
                    if len(detalhes) > 50:
                        detalhes = detalhes[:47] + '...'
                    
                    regras_data.append([nome, resultado, impacto, detalhes])
                
                tabela_regras = Table(regras_data, colWidths=[4*cm, 2.5*cm, 2.5*cm, 6*cm])
                tabela_regras.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (2, -1), 'CENTER'),
                    ('ALIGN', (3, 0), (3, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                elementos.append(tabela_regras)
        
        # Rodapé
        elementos.append(Spacer(1, 2*cm))
        rodape = Paragraph(
            f"<i>Relatório gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - IntegaGal v1.0</i>",
            ParagraphStyle('Footer', parent=estilo_normal, fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        )
        elementos.append(rodape)
        
        # Gerar PDF
        doc.build(elementos)
        
        return caminho_arquivo
    
    def exportar_exame_excel(self, dados_exame: Dict[str, Any], nome_arquivo: Optional[str] = None) -> str:
        """
        Exporta detalhes de um exame para Excel
        
        Args:
            dados_exame: Dicionário com dados do exame
            nome_arquivo: Nome do arquivo (opcional)
        
        Returns:
            Caminho completo do arquivo gerado
        """
        # Nome do arquivo
        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            exame_nome = dados_exame.get('exame', 'exame').replace(' ', '_')
            nome_arquivo = f"relatorio_exame_{exame_nome}_{timestamp}.xlsx"
        
        caminho_arquivo = os.path.join(self.pasta_saida, nome_arquivo)
        
        # Criar workbook
        wb = Workbook()
        
        # Estilos
        fonte_titulo = Font(name='Calibri', size=16, bold=True, color='1E88E5')
        fonte_subtitulo = Font(name='Calibri', size=12, bold=True, color='1E88E5')
        fonte_header = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        fonte_normal = Font(name='Calibri', size=10)
        
        preenchimento_header = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
        preenchimento_info = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        
        alinhamento_centro = Alignment(horizontal='center', vertical='center')
        alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
        
        borda = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Aba 1: Informações Gerais
        ws_info = wb.active
        ws_info.title = "Informações"
        
        # Título
        ws_info['A1'] = f"Relatório de Análise - {dados_exame.get('exame', 'N/A')}"
        ws_info['A1'].font = fonte_titulo
        ws_info.merge_cells('A1:D1')
        
        # Informações
        linha = 3
        ws_info[f'A{linha}'] = 'Data/Hora:'
        ws_info[f'A{linha}'].font = Font(bold=True)
        ws_info[f'A{linha}'].fill = preenchimento_info
        ws_info[f'B{linha}'] = dados_exame.get('data_hora', 'N/A')
        
        linha += 1
        ws_info[f'A{linha}'] = 'Equipamento:'
        ws_info[f'A{linha}'].font = Font(bold=True)
        ws_info[f'A{linha}'].fill = preenchimento_info
        ws_info[f'B{linha}'] = dados_exame.get('equipamento', 'N/A')
        
        linha += 1
        ws_info[f'A{linha}'] = 'Status:'
        ws_info[f'A{linha}'].font = Font(bold=True)
        ws_info[f'A{linha}'].fill = preenchimento_info
        ws_info[f'B{linha}'] = dados_exame.get('status', 'N/A').upper()
        
        if 'analista' in dados_exame:
            linha += 1
            ws_info[f'A{linha}'] = 'Analista:'
            ws_info[f'A{linha}'].font = Font(bold=True)
            ws_info[f'A{linha}'].fill = preenchimento_info
            ws_info[f'B{linha}'] = dados_exame['analista']
        
        # Ajustar larguras
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 40
        
        # Aba 2: Alvos
        alvos = dados_exame.get('alvos', {})
        if alvos:
            ws_alvos = wb.create_sheet(title="Alvos")
            
            # Headers
            headers = ['Alvo', 'CT', 'Resultado', 'Status']
            for col, header in enumerate(headers, start=1):
                cell = ws_alvos.cell(row=1, column=col, value=header)
                cell.font = fonte_header
                cell.fill = preenchimento_header
                cell.alignment = alinhamento_centro
                cell.border = borda
            
            # Dados
            for idx, (nome_alvo, dados) in enumerate(alvos.items(), start=2):
                ws_alvos.cell(row=idx, column=1, value=nome_alvo).border = borda
                
                ct = dados.get('ct')
                if isinstance(ct, (int, float)):
                    ws_alvos.cell(row=idx, column=2, value=ct).number_format = '0.00'
                else:
                    ws_alvos.cell(row=idx, column=2, value='N/D')
                ws_alvos.cell(row=idx, column=2).border = borda
                ws_alvos.cell(row=idx, column=2).alignment = alinhamento_centro
                
                resultado = dados.get('resultado', 'N/D')
                ws_alvos.cell(row=idx, column=3, value=resultado).border = borda
                
                status = 'OK' if resultado in ('Detectado', 'Positivo') else 'N/D'
                ws_alvos.cell(row=idx, column=4, value=status).border = borda
                ws_alvos.cell(row=idx, column=4).alignment = alinhamento_centro
            
            # Ajustar larguras
            ws_alvos.column_dimensions['A'].width = 20
            ws_alvos.column_dimensions['B'].width = 12
            ws_alvos.column_dimensions['C'].width = 20
            ws_alvos.column_dimensions['D'].width = 12
        
        # Aba 3: Controles
        controles = dados_exame.get('controles', {})
        if controles:
            ws_controles = wb.create_sheet(title="Controles")
            
            # Headers
            headers = ['Controle', 'Tipo', 'CT', 'Status']
            for col, header in enumerate(headers, start=1):
                cell = ws_controles.cell(row=1, column=col, value=header)
                cell.font = fonte_header
                cell.fill = preenchimento_header
                cell.alignment = alinhamento_centro
                cell.border = borda
            
            # Dados
            for idx, (nome_controle, dados) in enumerate(controles.items(), start=2):
                ws_controles.cell(row=idx, column=1, value=nome_controle).border = borda
                ws_controles.cell(row=idx, column=2, value=dados.get('tipo', 'N/A')).border = borda
                
                ct = dados.get('ct')
                if isinstance(ct, (int, float)):
                    ws_controles.cell(row=idx, column=3, value=ct).number_format = '0.00'
                else:
                    ws_controles.cell(row=idx, column=3, value='N/D')
                ws_controles.cell(row=idx, column=3).border = borda
                ws_controles.cell(row=idx, column=3).alignment = alinhamento_centro
                
                ws_controles.cell(row=idx, column=4, value=dados.get('status', 'N/A')).border = borda
                ws_controles.cell(row=idx, column=4).alignment = alinhamento_centro
            
            # Ajustar larguras
            ws_controles.column_dimensions['A'].width = 25
            ws_controles.column_dimensions['B'].width = 15
            ws_controles.column_dimensions['C'].width = 12
            ws_controles.column_dimensions['D'].width = 12
        
        # Aba 4: Regras
        regras_resultado = dados_exame.get('regras_resultado')
        if regras_resultado and regras_resultado.get('validacoes'):
            ws_regras = wb.create_sheet(title="Regras")
            
            # Headers
            headers = ['Regra', 'Resultado', 'Impacto', 'Detalhes']
            for col, header in enumerate(headers, start=1):
                cell = ws_regras.cell(row=1, column=col, value=header)
                cell.font = fonte_header
                cell.fill = preenchimento_header
                cell.alignment = alinhamento_centro
                cell.border = borda
            
            # Dados
            validacoes = regras_resultado.get('validacoes', [])
            for idx, validacao in enumerate(validacoes, start=2):
                ws_regras.cell(row=idx, column=1, value=validacao.get('regra_nome', 'N/A')).border = borda
                ws_regras.cell(row=idx, column=2, value=validacao.get('resultado', 'N/A')).border = borda
                ws_regras.cell(row=idx, column=2).alignment = alinhamento_centro
                ws_regras.cell(row=idx, column=3, value=validacao.get('impacto', 'N/A').upper()).border = borda
                ws_regras.cell(row=idx, column=3).alignment = alinhamento_centro
                ws_regras.cell(row=idx, column=4, value=validacao.get('detalhes', '')).border = borda
            
            # Ajustar larguras
            ws_regras.column_dimensions['A'].width = 30
            ws_regras.column_dimensions['B'].width = 15
            ws_regras.column_dimensions['C'].width = 12
            ws_regras.column_dimensions['D'].width = 50
        
        # Salvar
        wb.save(caminho_arquivo)
        
        return caminho_arquivo
    
    def exportar_historico_csv(self, df: pd.DataFrame, nome_arquivo: Optional[str] = None) -> str:
        """
        Exporta histórico de análises para CSV
        
        Args:
            df: DataFrame com histórico
            nome_arquivo: Nome do arquivo (opcional)
        
        Returns:
            Caminho completo do arquivo gerado
        """
        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"historico_analises_{timestamp}.csv"
        
        caminho_arquivo = os.path.join(self.pasta_saida, nome_arquivo)
        
        # Exportar
        df.to_csv(caminho_arquivo, index=False, encoding='utf-8-sig', sep=';')
        
        return caminho_arquivo
    
    def exportar_historico_excel(self, df: pd.DataFrame, nome_arquivo: Optional[str] = None) -> str:
        """
        Exporta histórico de análises para Excel com formatação
        
        Args:
            df: DataFrame com histórico
            nome_arquivo: Nome do arquivo (opcional)
        
        Returns:
            Caminho completo do arquivo gerado
        """
        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"historico_analises_{timestamp}.xlsx"
        
        caminho_arquivo = os.path.join(self.pasta_saida, nome_arquivo)
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico"
        
        # Estilos
        fonte_header = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        preenchimento_header = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
        alinhamento_centro = Alignment(horizontal='center', vertical='center')
        borda = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        for col, coluna in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col, value=coluna)
            cell.font = fonte_header
            cell.fill = preenchimento_header
            cell.alignment = alinhamento_centro
            cell.border = borda
        
        # Dados
        for idx, row in enumerate(df.itertuples(index=False), start=2):
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.border = borda
                if col > 1:  # Centralizar exceto primeira coluna
                    cell.alignment = alinhamento_centro
        
        # Ajustar larguras
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Salvar
        wb.save(caminho_arquivo)
        
        return caminho_arquivo


# Funções helper para uso direto
def exportar_pdf(dados_exame: Dict[str, Any], pasta_saida: str = "reports") -> str:
    """Atalho para exportar PDF"""
    exportador = ExportadorRelatorios(pasta_saida)
    return exportador.exportar_exame_pdf(dados_exame)


def exportar_excel(dados_exame: Dict[str, Any], pasta_saida: str = "reports") -> str:
    """Atalho para exportar Excel"""
    exportador = ExportadorRelatorios(pasta_saida)
    return exportador.exportar_exame_excel(dados_exame)


def exportar_csv(df: pd.DataFrame, pasta_saida: str = "reports") -> str:
    """Atalho para exportar CSV"""
    exportador = ExportadorRelatorios(pasta_saida)
    return exportador.exportar_historico_csv(df)


# Teste standalone
if __name__ == '__main__':
    # Dados de exemplo
    dados = {
        'exame': 'VR1e2 Biomanguinhos 7500 (Teste)',
        'data_hora': '08/12/2025 10:30:00',
        'equipamento': 'ABI 7500',
        'status': 'valida',
        'analista': 'Usuário Teste',
        'alvos': {
            'DEN1': {'ct': 18.5, 'resultado': 'Detectado'},
            'DEN2': {'ct': 22.3, 'resultado': 'Detectado'},
            'DEN3': {'ct': None, 'resultado': 'Não Detectado'},
        },
        'controles': {
            'Controle Positivo': {'tipo': 'Interno', 'ct': 20.5, 'status': 'OK'},
            'Controle Negativo': {'tipo': 'Interno', 'ct': None, 'status': 'OK'},
        },
        'regras_resultado': {
            'status': 'valida',
            'detalhes': '3 passou, 0 falhou',
            'validacoes': [
                {
                    'regra_nome': 'Controle Positivo OK',
                    'resultado': 'passou',
                    'detalhes': 'Controle positivo dentro do esperado (CT: 20.5)',
                    'impacto': 'critico'
                },
                {
                    'regra_nome': 'Fórmula: CT_DEN1 < 30',
                    'resultado': 'passou',
                    'detalhes': 'Resultado: True (tempo: 0.5ms)',
                    'impacto': 'alto'
                }
            ]
        }
    }
    
    exportador = ExportadorRelatorios()
    
    print("Testando exportações...")
    
    # PDF
    pdf_path = exportador.exportar_exame_pdf(dados)
    print(f"✅ PDF gerado: {pdf_path}")
    
    # Excel
    excel_path = exportador.exportar_exame_excel(dados)
    print(f"✅ Excel gerado: {excel_path}")
    
    # CSV (histórico)
    import pandas as pd
    df_teste = pd.DataFrame({
        'data_hora': ['08/12/2025 10:00', '08/12/2025 11:00'],
        'exame': ['Exame 1', 'Exame 2'],
        'equipamento': ['ABI 7500', 'QuantStudio'],
        'status': ['Válida', 'Aviso']
    })
    csv_path = exportador.exportar_historico_csv(df_teste)
    print(f"✅ CSV gerado: {csv_path}")
    
    print("\nTodas as exportações concluídas com sucesso!")
