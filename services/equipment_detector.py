"""
Equipment Detector - Fase 1
Detecta automaticamente o equipamento de PCR a partir de arquivo XLSX.
Retorna top-3 matches com confiança e estrutura detectada.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook


class XlrdSheetWrapper:
    """Wrapper para xlrd.sheet para compatibilidade com openpyxl."""
    
    def __init__(self, xlrd_sheet, xlrd_book):
        self.xlrd_sheet = xlrd_sheet
        self.xlrd_book = xlrd_book
        self.max_row = xlrd_sheet.nrows
        self.max_column = xlrd_sheet.ncols
        self.title = xlrd_sheet.name
    
    def cell(self, row, col):
        """Simula openpyxl cell() - retorna objeto com .value."""
        # xlrd usa índices 0-based, openpyxl usa 1-based
        row_idx = row - 1
        col_idx = col - 1
        
        if row_idx < 0 or row_idx >= self.max_row:
            return _XlrdCellWrapper(None)
        if col_idx < 0 or col_idx >= self.max_column:
            return _XlrdCellWrapper(None)
        
        try:
            cell_value = self.xlrd_sheet.cell_value(row_idx, col_idx)
            return _XlrdCellWrapper(cell_value)
        except Exception:
            return _XlrdCellWrapper(None)
    
    def close(self):
        """Compatibilidade - xlrd não precisa close explícito."""
        if hasattr(self.xlrd_book, 'release_resources'):
            self.xlrd_book.release_resources()


class _XlrdCellWrapper:
    """Wrapper para célula xlrd."""
    
    def __init__(self, value):
        self.value = value if value != '' else None


@dataclass
class EquipmentPattern:
    """Padrão de equipamento conhecido."""
    
    nome: str
    headers_esperados: List[str]  # Headers que devem estar presentes
    colunas_esperadas: Dict[str, int]  # Mapeamento coluna -> índice (0-based)
    linha_inicio_dados: int  # Linha onde começam os dados (1-based)
    keywords: Optional[List[str]] = None  # Keywords para detecção (opcional)
    validacoes: Optional[Dict[str, Any]] = None  # Validações específicas
    score_peso: Optional[Dict[str, float]] = None  # Peso de cada critério no score
    comentario: str = ""  # Comentário descritivo
    
    def __post_init__(self):
        """Inicializa valores padrão."""
        if self.validacoes is None:
            self.validacoes = {}
        if self.score_peso is None:
            self.score_peso = {'headers': 30, 'colunas': 25, 'linha_inicio': 15, 'validacoes': 30}
        if self.keywords is None:
            self.keywords = []


class EquipmentDetector:
    """Detector de equipamentos de PCR."""
    
    def __init__(self):
        self.padroes = obter_padroes_conhecidos()
    
    def detectar_equipamento(self, caminho_arquivo: str) -> Dict[str, Any]:
        """
        Detecta equipamento a partir de arquivo XLSX.
        
        Args:
            caminho_arquivo: Caminho para arquivo XLSX
            
        Returns:
            Dict com:
                - equipamento: Nome do equipamento detectado
                - confianca: Score de confiança (0-100)
                - alternativas: Lista de matches alternativos [{equipamento, confianca}]
                - estrutura_detectada: Estrutura do arquivo detectada
                
        Raises:
            FileNotFoundError: Se arquivo não existe
            ValueError: Se arquivo não é XLSX válido
        """
        # Validar arquivo
        path = Path(caminho_arquivo)
        if not path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")
        
        if path.suffix.lower() not in ['.xlsx', '.xls', '.xlsm']:
            raise ValueError(f"Arquivo deve ser XLSX/XLS/XLSM, recebido: {path.suffix}")
        
        try:
            # Analisar estrutura do arquivo (todas as abas)
            estrutura = analisar_estrutura_xlsx(caminho_arquivo)
            
            # Filtrar sheets de extração (devem ser ignoradas)
            if 'sheet_name' in estrutura:
                sheet_name_lower = estrutura['sheet_name'].lower()
                skip_keywords = ['extração', 'extracao', 'extraction']
                if any(kw in sheet_name_lower for kw in skip_keywords):
                    # Se for sheet de extração, tentar próxima aba
                    raise ValueError(f"Sheet '{estrutura['sheet_name']}' é de extração, ignorada.")
            
            # Calcular scores para cada padrão conhecido
            scores = []
            for padrao in self.padroes:
                score = calcular_match_score(estrutura, padrao)
                scores.append({
                    'equipamento': padrao.nome,
                    'confianca': score,
                    'padrao': padrao
                })
            
            # Ordenar por score (descrescente)
            scores.sort(key=lambda x: x['confianca'], reverse=True)
            
            # Pegar top-3
            melhor = scores[0]
            alternativas = [
                {'equipamento': s['equipamento'], 'confianca': s['confianca']}
                for s in scores[1:4]  # Top 2-4
            ]
            
            # Montar resultado
            resultado = {
                'equipamento': melhor['equipamento'],
                'confianca': melhor['confianca'],
                'alternativas': alternativas,
                'estrutura_detectada': {
                    'coluna_well': estrutura.get('coluna_well'),
                    'coluna_sample': estrutura.get('coluna_sample'),
                    'coluna_target': estrutura.get('coluna_target'),
                    'coluna_ct': estrutura.get('coluna_ct'),
                    'linha_inicio': estrutura.get('linha_inicio_dados'),
                    'headers': estrutura.get('headers', []),
                    'total_linhas': estrutura.get('total_linhas_dados'),
                    'total_colunas': estrutura.get('total_colunas')
                }
            }
            
            return resultado
            
        except Exception as e:
            raise ValueError(f"Erro ao analisar arquivo XLSX: {str(e)}") from e


def analisar_estrutura_xlsx(caminho_arquivo: str) -> Dict[str, Any]:
    """
    Analisa estrutura de arquivo XLSX para detecção de equipamento.
    
    Args:
        caminho_arquivo: Caminho para arquivo XLSX
        
    Returns:
        Dict com estrutura detectada:
            - headers: Lista de headers encontrados
            - colunas_nao_vazias: Índices de colunas com dados
            - linha_inicio_dados: Primeira linha com dados (1-based)
            - total_linhas_dados: Total de linhas com dados
            - total_colunas: Total de colunas
            - coluna_well: Índice da coluna de poços (se encontrada)
            - coluna_sample: Índice da coluna de amostras
            - coluna_target: Índice da coluna de alvos
            - coluna_ct: Índice da coluna de CT/Cq
            - amostras_wells: Exemplos de valores da coluna well
    """
    # Carregar com openpyxl para inspeção detalhada
    # Garantir leitura UTF-8 sem BOM
    path = Path(caminho_arquivo)
    
    # Escolher engine apropriado baseado na extensão
    if path.suffix.lower() == '.xls':
        # Para .xls, usar xlrd (suporte a formato Excel 97-2003)
        try:
            import xlrd
            
            # Abrir arquivo .xls com xlrd
            xls_book = xlrd.open_workbook(caminho_arquivo, on_demand=True)
            xls_sheet = xls_book.sheet_by_index(0)
            sheet_name = xls_sheet.name
            
            # Criar estrutura compatível com openpyxl
            # Usar wrapper customizado para xlrd
            wb = None
            ws = XlrdSheetWrapper(xls_sheet, xls_book)
            
        except ImportError:
            raise ImportError("Para ler arquivos .xls, instale: pip install xlrd")
        except Exception as e:
            raise ValueError(f"Erro ao ler arquivo .xls: {str(e)}")
    else:
        # Para .xlsx e .xlsm, usar openpyxl
        wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb.active
        sheet_name = ws.title if hasattr(ws, 'title') else "Sheet1"
    
    estrutura = {
        'sheet_name': sheet_name,  # Nome da aba para filtragem
        'headers': [],
        'colunas_nao_vazias': [],
        'linha_inicio_dados': None,
        'total_linhas_dados': 0,
        'total_colunas': ws.max_column,
        'coluna_well': None,
        'coluna_sample': None,
        'coluna_target': None,
        'coluna_ct': None,
        'amostras_wells': [],
        'conteudo_metadados': []  # Linhas 1-10 para detecção de keywords
    }
    
    # Procurar headers nas primeiras 30 linhas (alguns equipamentos colocam metadados extensos)
    headers_encontrados = False
    linha_header = None
    
    for row_idx in range(1, min(31, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            row_values.append(str(cell.value) if cell.value is not None else "")
        
        # Verificar se linha parece ser header
        # (tem strings não vazias em várias colunas)
        non_empty = [v for v in row_values if v.strip()]
        if len(non_empty) >= 3:
            # Verificar se tem palavras-chave típicas de header
            keywords = ['well', 'sample', 'target', 'cq', 'ct', 'name', 'bio-rad', 'c(t)']
            text_combined = " ".join(row_values).lower()
            
            if any(kw in text_combined for kw in keywords):
                headers_encontrados = True
                linha_header = row_idx
                estrutura['headers'] = row_values
                
                # CASO ESPECIAL: Formato CFX96_Export tem headers em 2 linhas
                # L1: Well, Name, Type... | L2: E gene, C(t), RdRP/S gene, C(t)...
                # Se encontrou Well na L1 mas não encontrou C(t), verificar L2
                if row_idx <= 2 and 'c(t)' not in text_combined.lower():
                    # Ler linha seguinte
                    next_row_idx = row_idx + 1
                    if next_row_idx <= ws.max_row:
                        next_row_values = []
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(next_row_idx, col_idx)
                            next_row_values.append(str(cell.value) if cell.value is not None else "")
                        
                        next_text = " ".join(next_row_values).lower()
                        if 'c(t)' in next_text:
                            # Merge: onde L1 está vazia, usar L2
                            merged_headers = []
                            for i, (h1, h2) in enumerate(zip(row_values, next_row_values)):
                                if h1.strip():
                                    merged_headers.append(h1)
                                elif h2.strip():
                                    merged_headers.append(h2)
                                else:
                                    merged_headers.append(f"Col{i+1}")
                            estrutura['headers'] = merged_headers
                            linha_header = row_idx  # Manter como L1, dados começam após L2
                
                break
    
    # Se não encontrou header explícito, assumir linha 1
    if not headers_encontrados:
        linha_header = 1
        estrutura['headers'] = [
            str(ws.cell(1, col).value or f"Col{col}")
            for col in range(1, ws.max_column + 1)
        ]
    
    # Identificar colunas específicas pelos headers
    headers_lower = [h.lower() for h in estrutura['headers']]
    
    for idx, header in enumerate(headers_lower):
        if any(kw in header for kw in ['well', 'poço', 'poco']):
            estrutura['coluna_well'] = idx
        elif any(kw in header for kw in ['sample', 'amostra']):
            estrutura['coluna_sample'] = idx
        elif any(kw in header for kw in ['target', 'alvo', 'detector']):
            estrutura['coluna_target'] = idx
        # CT e Cq são sinônimos (Cycle Threshold = Quantification Cycle)
        # Normalizar 'т' cirílico (U+0442) para 't' latino para detecção uniforme
        # Também reconhecer C(t) com parênteses (formato Bio-Rad CFX96 Export)
        elif any(kw in header for kw in ['cq', 'ct', 'cт', 'c т', 'c(t)']):
            # Priorizar colunas com nomes exatos (CT, Cт, Cq, C(t)) sem sufixos
            # Para evitar detectar "CT Threshold" em vez de "Cт"
            # IMPORTANTE: Pegar PRIMEIRA ocorrência quando há múltiplas (CFX96_Export tem várias colunas C(t))
            header_stripped = header.strip()
            if (header_stripped in ['cq', 'ct', 'cт', 'c(t)'] or len(header_stripped) <= 4) and estrutura.get('coluna_ct') is None:
                estrutura['coluna_ct'] = idx
        elif estrutura.get('coluna_ct') is None:
            # Fallback: aceitar com sufixos (threshold, mean, etc.) se nenhuma coluna exata foi encontrada
            if any(kw in header for kw in ['threshold', 'cycle', 'quantification']) and any(kw in header for kw in ['ct', 'cт', 'cq', 'c(t)']):
                estrutura['coluna_ct'] = idx
    
    # Procurar linha de início de dados (após header)
    linha_inicio = linha_header + 1
    for row_idx in range(linha_header + 1, min(linha_header + 10, ws.max_row + 1)):
        row_values = [
            str(ws.cell(row_idx, col).value or "")
            for col in range(1, ws.max_column + 1)
        ]
        
        non_empty = [v for v in row_values if v.strip()]
        if len(non_empty) >= 2:
            linha_inicio = row_idx
            estrutura['linha_inicio_dados'] = row_idx
            break
    
    if estrutura['linha_inicio_dados'] is None:
        estrutura['linha_inicio_dados'] = linha_header + 1
    
    # Identificar colunas não vazias
    colunas_com_dados = set()
    linhas_com_dados = 0
    
    for row_idx in range(estrutura['linha_inicio_dados'], min(ws.max_row + 1, estrutura['linha_inicio_dados'] + 100)):
        row_has_data = False
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value is not None and str(cell_value).strip():
                colunas_com_dados.add(col_idx - 1)  # 0-based
                row_has_data = True
        
        if row_has_data:
            linhas_com_dados += 1
    
    estrutura['colunas_nao_vazias'] = sorted(list(colunas_com_dados))
    estrutura['total_linhas_dados'] = linhas_com_dados
    
    # Coletar amostras de valores da coluna well (se identificada)
    if estrutura['coluna_well'] is not None:
        amostras = []
        for row_idx in range(estrutura['linha_inicio_dados'], min(ws.max_row + 1, estrutura['linha_inicio_dados'] + 10)):
            cell_value = ws.cell(row_idx, estrutura['coluna_well'] + 1).value
            if cell_value:
                amostras.append(str(cell_value))
        estrutura['amostras_wells'] = amostras
    
    # Coletar conteúdo das primeiras 10 linhas para detecção de keywords
    # (metadados geralmente ficam nessas linhas)
    metadados = []
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value is not None:
                row_values.append(str(cell_value))
        if row_values:
            metadados.append(" ".join(row_values))
    estrutura['conteudo_metadados'] = metadados
    
    # Fechar workbook (compatível com xlrd e openpyxl)
    if wb is not None:
        wb.close()
    elif hasattr(ws, 'close'):
        ws.close()
    
    return estrutura


def calcular_match_score(estrutura: Dict[str, Any], padrao: EquipmentPattern) -> float:
    """
    Calcula score de match entre estrutura detectada e padrão de equipamento.
    
    Args:
        estrutura: Estrutura detectada do arquivo
        padrao: Padrão de equipamento para comparar
        
    Returns:
        Score de 0 a 100
    """
    score = 0.0
    pesos = padrao.score_peso
    
    # 1. Verificar presença de headers esperados (peso: 30)
    headers_lower = [h.lower() for h in estrutura.get('headers', [])]
    headers_text = " ".join(headers_lower)
    
    headers_encontrados = 0
    for header_esperado in padrao.headers_esperados:
        if header_esperado.lower() in headers_text:
            headers_encontrados += 1
    
    if padrao.headers_esperados:
        score += (headers_encontrados / len(padrao.headers_esperados)) * pesos.get('headers', 30)
    
    # 2. Verificar colunas esperadas estão nas posições corretas (peso: 25)
    colunas_corretas = 0
    total_colunas_esperadas = len(padrao.colunas_esperadas)
    
    for nome_col, idx_esperado in padrao.colunas_esperadas.items():
        idx_detectado = estrutura.get(f'coluna_{nome_col}')
        if idx_detectado is not None:
            # Aceitar +/- 2 colunas de diferença
            if abs(idx_detectado - idx_esperado) <= 2:
                colunas_corretas += 1
    
    if total_colunas_esperadas > 0:
        score += (colunas_corretas / total_colunas_esperadas) * pesos.get('colunas', 25)
    
    # 3. Verificar linha de início de dados (peso: 15)
    linha_inicio_detectada = estrutura.get('linha_inicio_dados', 0)
    linha_inicio_esperada = padrao.linha_inicio_dados
    
    # Aceitar +/- 3 linhas de diferença
    if abs(linha_inicio_detectada - linha_inicio_esperada) <= 3:
        score += pesos.get('linha_inicio', 15)
    elif abs(linha_inicio_detectada - linha_inicio_esperada) <= 5:
        score += pesos.get('linha_inicio', 15) * 0.5
    
    # 4. Validações específicas (peso: 30)
    validacoes_ok = 0
    total_validacoes = 0  # Contar apenas validações aplicáveis
    
    for tipo_validacao, valor_esperado in padrao.validacoes.items():
        if tipo_validacao == 'skip_sheets':
            # Não é validação de match, skip
            continue
        
        total_validacoes += 1
        
        if tipo_validacao == 'formato_well':
            # Verificar formato de poços (A01, A1, etc.)
            amostras_wells = estrutura.get('amostras_wells', [])
            if amostras_wells:
                # Padrão: letra + número (A01, A1, A12, etc.)
                pattern = re.compile(r'^[A-H][0-9]{1,2}$', re.IGNORECASE)
                wells_validos = sum(1 for w in amostras_wells if pattern.match(str(w).strip()))
                if wells_validos / len(amostras_wells) >= 0.7:  # 70% válidos
                    validacoes_ok += 1
        
        elif tipo_validacao == 'min_linhas_dados':
            if estrutura.get('total_linhas_dados', 0) >= valor_esperado:
                validacoes_ok += 1
        
        elif tipo_validacao == 'keyword_presente':
            # Suporta string única ou lista de keywords
            keywords_to_check = valor_esperado if isinstance(valor_esperado, list) else [valor_esperado]
            
            # Buscar keywords nos metadados (linhas 1-10) e headers
            metadados_text = " ".join(estrutura.get('conteudo_metadados', [])).lower()
            headers_text = " ".join(estrutura.get('headers', [])).lower()
            combined_text = f"{metadados_text} {headers_text}"
            
            # Se qualquer keyword for encontrada, considera OK
            if any(kw.lower() in combined_text for kw in keywords_to_check):
                validacoes_ok += 1
        
        elif tipo_validacao == 'coluna_obrigatoria':
            if estrutura.get(f'coluna_{valor_esperado}') is not None:
                validacoes_ok += 1
    
    if total_validacoes > 0:
        score += (validacoes_ok / total_validacoes) * pesos.get('validacoes', 30)
    
    return round(score, 2)


def obter_padroes_conhecidos() -> List[EquipmentPattern]:
    """
    Retorna lista de padrões de equipamentos conhecidos.
    
    Returns:
        Lista de EquipmentPattern
    """
    padroes = []
    
    # 1. Applied Biosystems 7500 Real-Time PCR System
    padroes.append(EquipmentPattern(
        nome="7500",
        headers_esperados=["Well", "Sample Name", "Target", "Cq"],
        colunas_esperadas={
            'well': 0,      # Coluna A (índice 0)
            'sample': 1,    # Coluna B
            'target': 2,    # Coluna C
            'ct': 3         # Coluna D
        },
        linha_inicio_dados=5,
        validacoes={
            'formato_well': 'A01',
            'min_linhas_dados': 8,
            'coluna_obrigatoria': 'well',
            'coluna_obrigatoria_2': 'ct'
        },
        score_peso={
            'headers': 30,
            'colunas': 25,
            'linha_inicio': 15,
            'validacoes': 30
        }
    ))
    
    # 2. Bio-Rad CFX96 Touch Real-Time PCR Detection System
    padroes.append(EquipmentPattern(
        nome="CFX96",
        headers_esperados=["Well", "Fluor", "Target", "Sample", "Cq"],
        colunas_esperadas={
            'well': 0,      # Coluna A
            'target': 2,    # Coluna C (Target)
            'sample': 4,    # Coluna E (Sample)
            'ct': 5         # Coluna F (Cq)
        },
        linha_inicio_dados=21,  # Headers na linha 20, dados começam na 21
        keywords=["cfx", "bio-rad", "bio rad"],
        validacoes={
            'formato_well': 'A01',
            'keyword_presente': ['bio-rad', 'cfx', 'bio rad'],
            'min_linhas_dados': 50,
            'coluna_obrigatoria': 'well'
        },
        score_peso={
            'headers': 30,
            'colunas': 25,
            'linha_inicio': 15,
            'validacoes': 30
        },
        comentario="Bio-Rad CFX96 (headers na linha 20, Cq na coluna F)"
    ))
    
    # 2b. Bio-Rad CFX96 - Formato Export/Simplified (usa C(t) com parênteses)
    # Este formato tem DOIS níveis de headers:
    #   L1: Sample No, Patient Id, Well, Name, Type, FAM, Cal Red 610...
    #   L2: E gene, C(t), RdRP/S gene, C(t), N gene, C(t), IC, C(t)...
    padroes.append(EquipmentPattern(
        nome="CFX96_Export",
        headers_esperados=["Well", "Name", "Type", "E gene", "C(t)", "RdRP"],  # Mix L1+L2
        colunas_esperadas={
            'well': 2,      # Coluna C (Well) - linha 1
            'sample': 3,    # Coluna D (Name/Patient ID) - linha 1
            'target': 5,    # Coluna F (E gene, RdRP/S gene...) - linha 2, alvos múltiplos
            'ct': 6         # Coluna G - primeiro C(t) (E gene) - linha 2
        },
        linha_inicio_dados=3,  # Headers nas linhas 1-2, dados começam na 3
        keywords=["c(t)", "sample no", "patient id", "e gene", "rdrp"],
        validacoes={
            'formato_well': 'A01',
            'keyword_presente': ['c(t)'],  # Característica única: C(t) com parênteses na L2
            'min_linhas_dados': 50,
            'coluna_obrigatoria': 'well'
        },
        score_peso={
            'headers': 30,
            'colunas': 25,
            'linha_inicio': 15,
            'validacoes': 30
        },
        comentario="Bio-Rad CFX96 Export (2 níveis de headers L1-L2, C(t) com parênteses na L2)"
    ))
    
    # 3. Thermo Fisher QuantStudio 6 Pro System
    padroes.append(EquipmentPattern(
        nome="QuantStudio",
        headers_esperados=["Well", "Well Position", "Sample", "Target", "Cq"],
        colunas_esperadas={
            'well': 0,      # Coluna A (Well)
            'sample': 3,    # Coluna D (Sample)
            'target': 4,    # Coluna E (Target)
            'ct': 12        # Coluna M (Cq)
        },
        linha_inicio_dados=25,  # Headers na linha 24, dados começam na 25
        keywords=["quantstudio", "thermo fisher"],
        validacoes={
            'formato_well': 'A1',  # QuantStudio usa formato sem zero (A1, não A01)
            'min_linhas_dados': 50,
            'keyword_presente': ['quantstudio', 'thermo'],
            'coluna_obrigatoria': 'well'
        },
        score_peso={
            'headers': 30,
            'colunas': 25,
            'linha_inicio': 15,
            'validacoes': 30
        },
        comentario="Thermo Fisher QuantStudio (headers na linha 24, Cq na coluna M)"
    ))
    
    # 4. Applied Biosystems 7500 (variante com metadados estendidos)
    # Detectado por keywords "7500", "sds7500" ou "Applied Biosystems" no conteúdo da planilha
    padroes.append(EquipmentPattern(
        nome="7500_Extended",
        headers_esperados=["Well", "Sample Name", "Target Name", "Cт"],
        colunas_esperadas={
            'well': 0,      # Coluna A
            'sample': 1,    # Coluna B
            'target': 2,    # Coluna C
            'ct': 6         # Coluna G - Cт (valor real, não Mean)
        },
        linha_inicio_dados=9,  # Após linhas de metadados (linhas 1-7)
        keywords=["sds7500", "7500", "applied biosystems"],  # Keywords para detecção
        validacoes={
            'formato_well': 'A1',  # Formato sem zero à esquerda
            'min_linhas_dados': 50,  # Geralmente ~96 linhas
            'keyword_presente': ['sds7500', '7500', 'applied biosystems'],  # Lista de keywords
            'coluna_obrigatoria': 'well',
            'skip_sheets': ['extração', 'extracao', 'extraction']  # Ignorar abas com esses nomes
        },
        score_peso={
            'headers': 30,
            'colunas': 25,
            'linha_inicio': 15,
            'validacoes': 30
        }
    ))
    
    return padroes


# API pública
__all__ = [
    'EquipmentDetector',
    'EquipmentPattern',
    'detectar_equipamento',
    'analisar_estrutura_xlsx',
    'calcular_match_score',
    'obter_padroes_conhecidos'
]


def detectar_equipamento(caminho_arquivo: str) -> Dict[str, Any]:
    """
    Função helper para detectar equipamento.
    Wrapper para EquipmentDetector.detectar_equipamento.
    """
    detector = EquipmentDetector()
    return detector.detectar_equipamento(caminho_arquivo)
