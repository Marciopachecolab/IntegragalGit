"""
services.plate_viewer
---------------------

Módulo de visualização de placa 8x12 (96 poços físicos) para o Integragal.

Responsabilidades:
- Converter resultados normalizados da corrida (df_norm) + gabarito de extração (df_gabarito)
  em uma lista de WellResult.
- Exibir a placa em uma interface gráfica (customtkinter, com fallback para tkinter).
- Exportar a placa para Excel com formatação amigável.

Assinaturas principais:
    construir_well_results(df_norm, df_gabarito, config_regras) -> List[WellResult]
    mostrar_placa_gui(well_results, parent_window=None) -> None
    exportar_placa_excel(well_results, caminho_arquivo: str) -> None

df_norm esperado:
    - Uma linha por poço/target.
    - Colunas mínimas: 'well', 'target_name', 'ct'

df_gabarito esperado:
    - Deve possuir ao menos:
        * 'Poco' ou 'Poço' -> poço de análise (A1..H12)
        * 'Código' ou 'Codigo' ou 'Amostra' -> código da amostra/controle

config_regras esperado:
    - Dicionário contendo, idealmente:
        * 'alvos'                 -> string separada por ';' (ex.: 'SC2;RP')
        * 'CT_DETECTAVEL_MIN'     -> str/float
        * 'CT_DETECTAVEL_MAX'     -> str/float
        * 'CT_INCONCLUSIVO_MIN'   -> str/float
        * 'CT_INCONCLUSIVO_MAX'   -> str/float
        * 'CT_RP_MIN'             -> str/float
        * 'CT_RP_MAX'             -> str/float
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any

import math
import pandas as pd

# ---------------------------------------------------------------------------
# Tentativa de import do customtkinter; fallback para tkinter puro
# ---------------------------------------------------------------------------
try:
    import customtkinter as ctk  # type: ignore
    USE_CTK = True
except Exception:  # ImportError + outros ambientes sem GUI
    USE_CTK = False

# ---------------------------------------------------------------------------
# Excel (openpyxl)
# ---------------------------------------------------------------------------
try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore

    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

ROWS = "ABCDEFGH"  # 8 linhas físicas da placa


@dataclass
class WellResult:
    """Resultado consolidado por poço físico da placa."""
    well: str                   # ex: 'A1'
    sample_code: str            # código da amostra (ou CN/CP)
    ct_target: Optional[float]  # CT do alvo principal
    ct_rp: Optional[float]      # CT do RP consolidado da amostra
    result: str                 # 'Detectado', 'Nao Detectado', 'Inconclusivo', 'Invalido' ou ''
    is_control: bool = False    # True para CN/CP
    control_type: Optional[str] = None  # 'CN', 'CP', etc.


# ---------------------------------------------------------------------------
# Construção dos WellResult a partir dos DataFrames
# ---------------------------------------------------------------------------


def construir_well_results(
    df_norm: pd.DataFrame,
    df_gabarito: pd.DataFrame,
    config_regras: Dict[str, Any],
) -> List[WellResult]:
    """
    Constrói uma lista de WellResult a partir de:
    - df_norm: resultados normalizados da corrida (uma linha por poço/target).
    - df_gabarito: gabarito de extração (poço -> amostra).
    - config_regras: dicionário com parâmetros do exame.
    """
    if df_norm is None or df_norm.empty:
        return []

    if df_gabarito is None or df_gabarito.empty:
        df_gab = pd.DataFrame(columns=["well", "sample_code"])
    else:
        df_gab = df_gabarito.copy()

    # ----------------------------
    # 1) Normalização do gabarito
    # ----------------------------
    col_poco = None
    for cand in ["Poco", "Poço", "POCO", "POÇO"]:
        if cand in df_gab.columns:
            col_poco = cand
            break
    if col_poco is None:
        raise ValueError("df_gabarito não possui coluna 'Poco' ou 'Poço'.")

    col_sample = None
    for cand in ["Código", "Codigo", "codigo", "codigo_amostra", "Amostra", "amostra"]:
        if cand in df_gab.columns:
            col_sample = cand
            break
    if col_sample is None:
        df_gab["sample_code"] = ""
    else:
        df_gab["sample_code"] = df_gab[col_sample].astype(str)

    df_gab["well"] = df_gab[col_poco].astype(str).str.upper().str.strip()
    df_gab_slim = df_gab[["well", "sample_code"]].drop_duplicates()

    # ----------------------------
    # 2) Normalização de df_norm
    # ----------------------------
    df_raw = df_norm.copy()
    if "well" not in df_raw.columns or "target_name" not in df_raw.columns:
        raise ValueError("df_norm precisa ter colunas 'well' e 'target_name'.")

    df_raw["well"] = df_raw["well"].astype(str).str.upper().str.strip()
    if "ct" not in df_raw.columns:
        if "ct_raw" in df_raw.columns:
            df_raw["ct"] = _normalizar_ct(df_raw["ct_raw"])
        else:
            raise ValueError("df_norm precisa ter coluna 'ct' ou 'ct_raw'.")

    # Junta gabarito para obter sample_code em cada linha
    df_raw = df_raw.merge(df_gab_slim, on="well", how="left")

    # ----------------------------
    # 3) Leitura de parâmetros de regras
    # ----------------------------
    alvos_str = str(config_regras.get("alvos") or "")
    alvos = [a.strip() for a in alvos_str.split(";") if a.strip()]
    alvo_principal = alvos[0] if alvos else None

    def _f(key: str, default: float) -> float:
        val = config_regras.get(key)
        if val is None or val == "":
            return default
        try:
            return float(str(val).replace(",", "."))
        except Exception:
            return default

    ct_detect_min = _f("CT_DETECTAVEL_MIN", 0.0)
    ct_detect_max = _f("CT_DETECTAVEL_MAX", 40.0)
    ct_inconc_min = _f("CT_INCONCLUSIVO_MIN", 40.01)
    ct_inconc_max = _f("CT_INCONCLUSIVO_MAX", 45.0)
    ct_rp_min = _f("CT_RP_MIN", 0.0)
    ct_rp_max = _f("CT_RP_MAX", 45.0)

    rp_names = ["RP", "RP_1", "RP_2"]

    # ----------------------------
    # 4) RP consolidado por amostra
    # ----------------------------
    df_rp = df_raw[
        df_raw["target_name"].astype(str).str.upper().isin([n.upper() for n in rp_names])
    ]
    rp_por_amostra: Dict[str, Optional[float]] = {}
    if not df_rp.empty:
        for sample_code, sub in df_rp.groupby("sample_code"):
            vals = [v for v in sub["ct"].tolist() if _is_num(v)]
            if vals:
                rp_por_amostra[sample_code] = float(sum(vals) / len(vals))
            else:
                rp_por_amostra[sample_code] = None

    # ----------------------------
    # 5) CT do alvo principal por poço
    # ----------------------------
    if alvo_principal:
        df_alvo = df_raw[
            df_raw["target_name"].astype(str).str.upper() == alvo_principal.upper()
        ]
    else:
        df_alvo = df_raw.iloc[0:0]

    ct_alvo_por_poco: Dict[str, Optional[float]] = {}
    if not df_alvo.empty:
        for well, sub in df_alvo.groupby("well"):
            vals = [v for v in sub["ct"].tolist() if _is_num(v)]
            ct_alvo_por_poco[well] = vals[0] if vals else None

    # ----------------------------
    # 6) Construção da matriz de 8x12 poços
    # ----------------------------
    resultados: List[WellResult] = []

    for r in range(8):  # linhas A..H
        for c in range(12):  # colunas 1..12
            well_name = f"{ROWS[r]}{c+1}"
            linha_gab = df_gab_slim[df_gab_slim["well"] == well_name]
            if not linha_gab.empty:
                sample_code = str(linha_gab["sample_code"].iloc[0])
            else:
                sample_code = ""

            ct_t = ct_alvo_por_poco.get(well_name)
            ct_rp = rp_por_amostra.get(sample_code)

            ctrl_type = _tipo_controle(sample_code)
            is_ctrl = ctrl_type is not None

            if alvo_principal and (sample_code or _is_num(ct_t) or _is_num(ct_rp)):
                result = _interpretar_com_rp(
                    ct_rp=ct_rp,
                    ct_alvo=ct_t,
                    ct_detect_min=ct_detect_min,
                    ct_detect_max=ct_detect_max,
                    ct_inconc_min=ct_inconc_min,
                    ct_inconc_max=ct_inconc_max,
                    ct_rp_min=ct_rp_min,
                    ct_rp_max=ct_rp_max,
                )
            else:
                result = ""

            resultados.append(
                WellResult(
                    well=well_name,
                    sample_code=sample_code,
                    ct_target=ct_t,
                    ct_rp=ct_rp,
                    result=result,
                    is_control=is_ctrl,
                    control_type=ctrl_type,
                )
            )

    return resultados


# ---------------------------------------------------------------------------
# Funções auxiliares de interpretação
# ---------------------------------------------------------------------------


def _is_num(x: Any) -> bool:
    try:
        if x is None or (isinstance(x, float) and math.isnan(x)):
            return False
        float(x)
        return True
    except Exception:
        return False


def _normalizar_ct(series_ct: pd.Series) -> pd.Series:
    """
    Converte valores especiais (UNDETERMINED, NA, etc.) para None/NaN
    e valores numéricos para float.
    """

    def conv(x: Any) -> Any:
        if x is None:
            return None
        s = str(x).strip().upper()
        if s in ("", "NA", "N/A", "UNDETERMINED", "UND", "ND"):
            return None
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return None

    return series_ct.apply(conv)


def _tipo_controle(sample_code: str) -> Optional[str]:
    s = (sample_code or "").upper()
    if "CN" in s:
        return "CN"
    if "CP" in s:
        return "CP"
    return None


def _interpretar_com_rp(
    ct_rp: Any,
    ct_alvo: Any,
    ct_detect_min: float,
    ct_detect_max: float,
    ct_inconc_min: float,
    ct_inconc_max: float,
    ct_rp_min: float,
    ct_rp_max: float,
) -> str:
    """
    Interpreta o resultado de um alvo considerando o RP da amostra.
    """
    if ct_rp is None:
        return "Invalido"
    try:
        valor_rp = float(ct_rp)
    except Exception:
        return "Invalido"

    if not (ct_rp_min <= valor_rp <= ct_rp_max):
        return "Invalido"

    if ct_alvo is None:
        return "Nao Detectado"

    try:
        valor_ct = float(ct_alvo)
    except Exception:
        return "Nao Detectado"

    if valor_ct <= ct_detect_max:
        return "Detectado"

    if ct_inconc_min <= valor_ct <= ct_inconc_max:
        return "Inconclusivo"

    return "Nao Detectado"


# ---------------------------------------------------------------------------
# Conversão lista -> matriz de 8x12 (para GUI/Excel)
# ---------------------------------------------------------------------------


def _results_to_matrix(results: List[WellResult]) -> Dict[Tuple[int, int], WellResult]:
    """Indexa WellResult por posição (linha, coluna) 0-based."""
    idx: Dict[str, WellResult] = {r.well.upper(): r for r in results}
    matrix: Dict[Tuple[int, int], WellResult] = {}
    for r in range(8):
        for c in range(12):
            well = f"{ROWS[r]}{c+1}".upper()
            matrix[(r, c)] = idx.get(
                well,
                WellResult(
                    well=well,
                    sample_code="",
                    ct_target=None,
                    ct_rp=None,
                    result="",
                    is_control=False,
                    control_type=None,
                ),
            )
    return matrix


# ---------------------------------------------------------------------------
# Interface gráfica
# ---------------------------------------------------------------------------


def mostrar_placa_gui(results: List[WellResult], parent_window: Any | None = None) -> None:
    """
    Abre uma janela exibindo a placa 8x12.
    """
    matrix = _results_to_matrix(results)

    if USE_CTK:
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        if parent_window is None:
            root = ctk.CTk()
        else:
            root = ctk.CTkToplevel(parent_window)
        root.title("Visualização de Placa")
        root.geometry("1120x600")
    else:
        import tkinter as tk  # type: ignore

        if parent_window is None:
            root = tk.Tk()
        else:
            root = tk.Toplevel(parent_window)
        root.title("Visualização de Placa")
        root.geometry("1120x600")

    # Cabeçalhos de coluna
    for c in range(12):
        label = _make_label(root, text=str(c + 1), bold=True)
        label.grid(row=0, column=c + 1, padx=2, pady=2, sticky="nsew")

    # Cabeçalhos de linha + células
    for r in range(8):
        row_label = _make_label(root, text=ROWS[r], bold=True)
        row_label.grid(row=r + 1, column=0, padx=2, pady=2, sticky="nsew")

        for c in range(12):
            well_res = matrix[(r, c)]
            frame = _make_well_frame(root, well_res)
            frame.grid(row=r + 1, column=c + 1, padx=2, pady=2, sticky="nsew")

    for c in range(13):
        root.grid_columnconfigure(c, weight=1)
    for r in range(9):
        root.grid_rowconfigure(r, weight=1)

    if parent_window is None:
        root.mainloop()


def _make_label(parent: Any, text: str, bold: bool = False):
    """Label para cabeçalhos, compatível com CTk ou Tk."""
    if USE_CTK:
        font = ("Arial", 11, "bold" if bold else "normal")
        return ctk.CTkLabel(parent, text=text, font=font)  # type: ignore[name-defined]
    else:
        import tkinter as tk  # type: ignore

        fnt = ("Arial", 11, "bold" if bold else "normal")
        return tk.Label(parent, text=text, font=fnt)


def _make_well_frame(parent: Any, well: WellResult):
    """Cria um quadrado de poço com informações de resultado."""
    bg_main = "white"
    side_bg, side_fg = "#DDDDDD", "black"

    if well.result == "Detectado":
        side_bg, side_fg = "black", "white"
    elif well.result == "Nao Detectado":
        side_bg, side_fg = "#DDDDDD", "black"
    elif well.result == "Inconclusivo":
        side_bg, side_fg = "orange", "black"
    elif well.result == "Invalido":
        side_bg, side_fg = "red", "white"

    if well.is_control:
        bg_main = "#E0F2FF"

    if USE_CTK:
        frame = ctk.CTkFrame(parent, fg_color=bg_main, corner_radius=4)  # type: ignore[name-defined]
        inner_font = ("Arial", 9)
        small_font = ("Arial", 8)
        side = ctk.CTkFrame(frame, width=10, fg_color=side_bg, corner_radius=2)  # type: ignore[name-defined]
        side.pack(side="left", fill="y")

        side_label = ctk.CTkLabel(  # type: ignore[name-defined]
            side,
            text=_short_result(well.result),
            font=small_font,
            text_color=side_fg,
        )
        side_label.pack(expand=True)

        code = ctk.CTkLabel(frame, text=well.sample_code, font=inner_font)  # type: ignore[name-defined]
        code.pack(side="top", fill="x", padx=2, pady=(2, 1))

        ct_text = f"CT: {well.ct_target if well.ct_target is not None else '--'}"
        rp_text = f"RP: {well.ct_rp if well.ct_rp is not None else '--'}"

        ct_label = ctk.CTkLabel(frame, text=ct_text, font=small_font)  # type: ignore[name-defined]
        ct_label.pack(side="top", fill="x")

        rp_label = ctk.CTkLabel(frame, text=rp_text, font=small_font)  # type: ignore[name-defined]
        rp_label.pack(side="top", fill="x", pady=(0, 2))

    else:
        import tkinter as tk  # type: ignore

        frame = tk.Frame(parent, bg=bg_main, bd=1, relief="solid")
        inner_font = ("Arial", 8)
        small_font = ("Arial", 7)

        side = tk.Frame(frame, width=10, bg=side_bg)
        side.pack(side="left", fill="y")

        side_label = tk.Label(
            side,
            text=_short_result(well.result),
            font=small_font,
            fg=side_fg,
            bg=side_bg,
        )
        side_label.pack(expand=True)

        code = tk.Label(frame, text=well.sample_code, font=inner_font, bg=bg_main)
        code.pack(side="top", fill="x", padx=2, pady=(2, 1))

        ct_text = f"CT: {well.ct_target if well.ct_target is not None else '--'}"
        rp_text = f"RP: {well.ct_rp if well.ct_rp is not None else '--'}"

        ct_label = tk.Label(frame, text=ct_text, font=small_font, bg=bg_main)
        ct_label.pack(side="top", fill="x")

        rp_label = tk.Label(frame, text=rp_text, font=small_font, bg=bg_main)
        rp_label.pack(side="top", fill="x", pady=(0, 2))

    return frame


def _short_result(result: str) -> str:
    """Converte resultado em abreviação para a barrinha lateral."""
    mapping = {
        "Detectado": "DET",
        "Nao Detectado": "ND",
        "Inconclusivo": "INC",
        "Invalido": "INV",
        "": "",
    }
    return mapping.get(result, result[:3].upper() if result else "")


# ---------------------------------------------------------------------------
# Exportação para Excel
# ---------------------------------------------------------------------------


def exportar_placa_excel(results: List[WellResult], caminho_arquivo: str | Path) -> None:
    """
    Gera um arquivo Excel com a placa 8x12.
    """
    if not HAS_OPENPYXL:
        print("openpyxl não está instalado; exportação Excel ignorada.")
        return

    matrix = _results_to_matrix(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "Placa"

    header_font = Font(bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Cores
    fill_det = PatternFill("solid", fgColor="FF9999")   # vermelho claro
    fill_nd = PatternFill("solid", fgColor="DDDDDD")    # cinza claro
    fill_inc = PatternFill("solid", fgColor="ADD8E6")   # azul claro
    fill_inv = PatternFill("solid", fgColor="C6EFCE")   # verde claro
    fill_ctrl = PatternFill("solid", fgColor="E0F2FF")  # azul claro para controles

    # Cabeçalhos
    ws["A1"] = ""
    for c in range(12):
        cell = ws.cell(row=1, column=c + 2)
        cell.value = c + 1
        cell.font = header_font
        cell.alignment = center

    for r in range(8):
        cell = ws.cell(row=r + 2, column=1)
        cell.value = ROWS[r]
        cell.font = header_font
        cell.alignment = center

    # Células de poços
    for r in range(8):
        for c in range(12):
            well = matrix[(r, c)]
            row_excel = r + 2
            col_excel = c + 2

            cell = ws.cell(row=row_excel, column=col_excel)

            lines = [
                well.sample_code or "",
                f"CT: {well.ct_target if well.ct_target is not None else '--'}",
                f"RP: {well.ct_rp if well.ct_rp is not None else '--'}",
                _short_result(well.result),
            ]
            cell.value = "\n".join(lines)
            cell.alignment = center
            cell.border = border_thin
            cell.font = Font(size=9, bold=True)

            fill = None
            if well.result == "Detectado":
                fill = fill_det
            elif well.result == "Nao Detectado":
                fill = fill_nd
            elif well.result == "Inconclusivo":
                fill = fill_inc
            elif well.result == "Invalido":
                fill = fill_inv

            if fill is not None:
                cell.fill = fill

            if well.is_control:
                cell.fill = fill_ctrl

    ws.column_dimensions["A"].width = 4
    for col_letter in "BCDEFGHIJKLM":
        ws.column_dimensions[col_letter].width = 14

    for row_idx in range(2, 10):
        ws.row_dimensions[row_idx].height = 45

    caminho = Path(caminho_arquivo)
    wb.save(caminho)
    print(f"Placa exportada para Excel em: {caminho.resolve()}")
