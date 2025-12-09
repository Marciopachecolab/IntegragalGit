import openpyxl

arquivo = r'C:\Users\marci\Downloads\18 JULHO 2025\teste\exemploseegene.xlsx'
wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
ws = wb.active

print("PRIMEIRAS 5 LINHAS COMPLETAS:")
for linha in range(1, 6):
    print(f"\nLinha {linha}:")
    for col in range(1, min(15, ws.max_column + 1)):
        val = ws.cell(linha, col).value
        if val:
            print(f"   Col {col} ({chr(64+col)}): {val}")

wb.close()
