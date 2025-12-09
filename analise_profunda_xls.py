"""
Análise profunda dos arquivos - lendo TODAS as linhas para encontrar Cq.
Baseado na imagem: equipamento em linha 5 coluna B, Cq em linha 24 colunas M:P
"""

from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent))

# Arquivos da imagem
arquivos = [
    r"C:\Users\marci\Downloads\18 JULHO 2025\teste\20210809 COVID BIO M PLACA 8_Copy_20210809_182622_Results_20210809 202116.xls",
    r"C:\Users\marci\Downloads\18 JULHO 2025\teste\SC2 20200729-MANAGER.xls",
    r"C:\Users\marci\Downloads\18 JULHO 2025\teste\20210809 COVID BIO M PLACA 8_Copy_20210809_182622_Resytryrtyrtyreytults_20210809 2021adadfadasdasd16.xls",
    r"C:\Users\marci\Downloads\18 JULHO 2025\teste\SC2 20200729-MANAGER.xlsx"
]

print("\n" + "="*80)
print("ANÁLISE PROFUNDA - LEITURA COMPLETA DOS ARQUIVOS")
print("="*80)

for i, arquivo_path in enumerate(arquivos, 1):
    arquivo = Path(arquivo_path)
    
    if not arquivo.exists():
        print(f"\n[{i}/4] ❌ Arquivo não encontrado: {arquivo.name}")
        continue
    
    print(f"\n{'='*80}")
    print(f"[{i}/4] 📂 {arquivo.name}")
    print(f"{'='*80}")
    print(f"Tamanho: {arquivo.stat().st_size / 1024:.1f} KB")
    
    try:
        # Ler com xlrd ou openpyxl
        if arquivo.suffix.lower() == '.xls':
            import xlrd
            wb = xlrd.open_workbook(str(arquivo))
            ws = wb.sheet_by_index(0)
            print(f"Sheet: '{ws.name}'")
            print(f"Dimensões: {ws.nrows} linhas x {ws.ncols} colunas")
            
            # Linha 5 coluna B (índice 1) - equipamento
            if ws.nrows >= 5:
                equipamento_cell = ws.cell_value(4, 1)  # Linha 5 = índice 4, coluna B = índice 1
                print(f"\n🔬 LINHA 5 COLUNA B (Equipamento):")
                print(f"   '{equipamento_cell}'")
                
                # Verificar se é QuantStudio ou CFX
                equip_lower = str(equipamento_cell).lower()
                if 'quantstudio' in equip_lower:
                    print(f"   ✅ QuantStudio detectado!")
                elif 'cfx' in equip_lower:
                    print(f"   ✅ CFX detectado!")
                elif '7500' in equip_lower or 'applied' in equip_lower:
                    print(f"   ✅ 7500 detectado!")
            
            # Linha 24 - procurar headers com Cq
            print(f"\n📋 LINHA 24 (Headers de dados):")
            if ws.nrows >= 24:
                linha_24 = []
                for col_idx in range(min(ws.ncols, 30)):  # Primeiras 30 colunas
                    cell_value = ws.cell_value(23, col_idx)  # Linha 24 = índice 23
                    if cell_value:
                        linha_24.append(f"[{col_idx}]{cell_value}")
                
                print(f"   {' | '.join(linha_24[:20])}")
                
                # Procurar por Cq/CT
                for col_idx in range(ws.ncols):
                    cell_value = str(ws.cell_value(23, col_idx)).lower()
                    if 'cq' in cell_value or 'ct' in cell_value:
                        original_value = ws.cell_value(23, col_idx)
                        print(f"   ✅ Coluna {col_idx} ({chr(65+col_idx)}): '{original_value}' contém Cq/CT")
            
            # Procurar em TODAS as linhas por headers com Well, Sample, Target, Cq
            print(f"\n🔍 BUSCA POR HEADERS (Well, Sample, Target, Cq):")
            headers_encontrados = []
            
            for row_idx in range(min(ws.nrows, 30)):  # Primeiras 30 linhas
                row_values = []
                for col_idx in range(min(ws.ncols, 20)):
                    cell_value = ws.cell_value(row_idx, col_idx)
                    if cell_value:
                        row_values.append(str(cell_value))
                
                row_text = " ".join(row_values).lower()
                
                # Verificar se tem palavras-chave de header
                keywords = ['well', 'sample', 'target', 'cq', 'ct', 'reporter', 'quencher']
                matches = [kw for kw in keywords if kw in row_text]
                
                if len(matches) >= 3:  # Pelo menos 3 keywords
                    print(f"   📍 Linha {row_idx + 1}: {matches}")
                    headers_encontrados.append({
                        'linha': row_idx + 1,
                        'keywords': matches,
                        'conteudo': ' | '.join(row_values[:10])
                    })
            
            if headers_encontrados:
                print(f"\n   ✅ {len(headers_encontrados)} linha(s) com headers encontrada(s):")
                for h in headers_encontrados:
                    print(f"      Linha {h['linha']}: {h['keywords']}")
                    print(f"         {h['conteudo'][:80]}...")
            
            # Amostras de dados (linhas 25-30)
            print(f"\n📊 AMOSTRAS DE DADOS (linhas 25-30):")
            if ws.nrows >= 25:
                for row_idx in range(24, min(ws.nrows, 30)):
                    row_sample = []
                    for col_idx in range(min(ws.ncols, 15)):
                        cell_value = ws.cell_value(row_idx, col_idx)
                        if cell_value:
                            row_sample.append(str(cell_value)[:20])
                    
                    if row_sample:
                        print(f"   L{row_idx + 1}: {' | '.join(row_sample[:8])}")
        
        else:  # .xlsx
            from openpyxl import load_workbook
            
            try:
                wb = load_workbook(str(arquivo), read_only=True, data_only=True)
                ws = wb.active
                print(f"Sheet: '{ws.title}'")
                print(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
                
                # Linha 5 coluna B
                if ws.max_row >= 5:
                    equipamento_cell = ws.cell(5, 2).value  # Linha 5, coluna B (índice 2)
                    print(f"\n🔬 LINHA 5 COLUNA B:")
                    print(f"   '{equipamento_cell}'")
                
                # Linha 24
                print(f"\n📋 LINHA 24:")
                if ws.max_row >= 24:
                    linha_24 = []
                    for col_idx in range(1, min(ws.max_column + 1, 30)):
                        cell_value = ws.cell(24, col_idx).value
                        if cell_value:
                            col_letter = chr(64 + col_idx)
                            linha_24.append(f"{col_letter}:{cell_value}")
                    
                    print(f"   {' | '.join(linha_24[:15])}")
                
                wb.close()
                
            except Exception as e:
                print(f"   ⚠️ Erro ao ler .xlsx: {str(e)[:100]}")
        
        print(f"\n{'='*80}")
        
    except Exception as e:
        print(f"\n❌ ERRO: {type(e).__name__}: {str(e)[:100]}")
        import traceback
        print(traceback.format_exc()[:500])

print("\n" + "="*80)
print("✅ ANÁLISE COMPLETA")
print("="*80)
