import openpyxl
import xlrd
from pathlib import Path

arquivo = r'C:\Users\marci\Downloads\18 JULHO 2025\teste\exemploseegene.xlsx'

print("="*80)
print(f"ANÁLISE: {Path(arquivo).name}")
print("="*80)

try:
    wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.active
    
    print(f"\n Sheet: '{ws.title}'")
    print(f" Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
    
    # Procurar por C(t) nas primeiras 30 linhas
    print("\n PROCURANDO C(t)...")
    ct_encontrado = False
    
    for linha_idx in range(1, min(31, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(linha_idx, col_idx).value
            if cell_value and 'C(t)' in str(cell_value):
                print(f"    ENCONTRADO na linha {linha_idx}, coluna {col_idx} ({chr(64+col_idx)})")
                print(f"      Valor: '{cell_value}'")
                ct_encontrado = True
    
    if not ct_encontrado:
        print("    C(t) não encontrado nas primeiras 30 linhas")
    
    # Mostrar headers das primeiras 25 linhas
    print("\n HEADERS (primeiras 25 linhas):")
    for linha_idx in range(1, min(26, ws.max_row + 1)):
        linha_valores = []
        tem_conteudo = False
        for col_idx in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(linha_idx, col_idx).value
            if val:
                tem_conteudo = True
                linha_valores.append(f"{chr(64+col_idx)}: {str(val)[:40]}")
        
        if tem_conteudo:
            print(f"   L{linha_idx:2d}: {' | '.join(linha_valores)}")
    
    # Buscar por keywords do Bio-Rad
    print("\n KEYWORDS (primeiras 10 linhas):")
    keywords = ['bio-rad', 'cfx', 'manager', 'c(t)']
    for linha_idx in range(1, min(11, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(linha_idx, col_idx).value or '').lower()
            for kw in keywords:
                if kw in cell_value:
                    print(f"    '{kw}' encontrado em L{linha_idx}:{chr(64+col_idx)}")
    
    wb.close()
    
except Exception as e:
    print(f" ERRO: {e}")
    import traceback
    traceback.print_exc()
